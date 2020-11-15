from math import ceil

import numpy as np
from sklearn.preprocessing import normalize
from sklearn.metrics.pairwise import cosine_similarity

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class Phenotype(object):
    def __init__(self, phenotypes, dim_names, stats):
        self.phenotypes = phenotypes
        self.dim_names = dim_names
        self.stats = stats

    def __getitem__(self, idx):
        return self.phenotypes[idx]

    def __len__(self):
        return len(self.phenotypes)


def interpret_phenotypes(factors, item_idx2desc, dim_names, threshold=None):
    """Function to interpret the non-negative CP factors as phenotypes. The columns of the CP factors
    are first normalized by their l1-norm; the values less than the threshold are then filtered out.

    Args:

        factors (iterable of np.ndarray): CP factor matrices, one matrix for each dimension.
        item_idx2desc (iterable of dict): Concept mapping of the factor matrices. Each one must be
            a {idx: description} dict, and must has the same order with factors.
        dim_names: The name of each dimensions.
        threshold: The hard threshold to filter out items with very small values.

    Returns:
        Phenotype: The phenotype interpretation of the input factors.
    """
    assert len(factors) == len(item_idx2desc), 'Number of factors and the concept mappings must be the same.'
    assert all([factor.min() >= 0 for factor in factors]), 'All CP factors must be non-negative.'

    phenotypes = []
    n_dims = len(factors)
    n_factors = factors[0].shape[1]
    item_sortidx = [np.argsort(-U, axis=0) for U in factors]

    factors = [normalize(factor, axis=0, norm='l1') for factor in factors]

    if threshold is not None:
        for i, factor in enumerate(factors):
            factor[factor < threshold[i]] = 0

    # get stats of phenotype overlapping
    def overlapping_stat(factor):
        positive_flag = factor > 0
        return positive_flag.sum(axis=1)

    stats = {
        'overlap': [overlapping_stat(factor) for factor in factors]
    }

    for r in range(n_factors):
        phenotype_definition = []
        for j in range(n_dims):
            dim_j = []
            for idx in item_sortidx[j][:, r]:
                if factors[j][idx, r] > 0:
                    dim_j.append((idx, item_idx2desc[j][idx], factors[j][idx, r]))
                else:
                    break
            phenotype_definition.append(dim_j)
        phenotypes.append(phenotype_definition)
    return Phenotype(phenotypes, dim_names, stats)


def sparsity_similarity(factors, threshold=1e-5):
    sparsity = [U[U > threshold].shape[0] / (U.shape[0] * U.shape[1]) for U in factors]
    similarity = []
    for U in factors:
        similarity_matrix = cosine_similarity(U.T)
        r = np.arange(similarity_matrix.shape[0])
        mask = r[:, None] < r
        similarity.append(similarity_matrix[mask].mean())
    return sparsity, similarity


def interpret_corres_matrix(corrs, row_mapping, col_mapping, col_normalize=True):
    """Function to interpret the correspondence matrix. The correspondence are interpreted column-wisely. If
    the correspondence matrix should be interpreted row-wisely, it should be transposed before calling this
    function.

    Args:
        corrs (np.ndarray): The correspondence matrix to be interpreted.
        row_mapping (dict): The idx to name mapping of each row.
        col_mapping (dict): The idx to name mapping of each column.
        col_normalize (bool): Indicate whether to normalize the columns, default is True.

    Returns:
        The correspondence of the row concepts to each column concept.
    """

    if col_normalize:
        corrs = normalize(corrs, axis=0, norm='l1')
    sorts = np.argsort(-corrs, axis=0)
    col_names = np.array([row_mapping[i] for i in range(len(row_mapping))])
    cols = []
    for i in range(corrs.shape[1]):
        col = list(map(lambda x: x if x[1] > 0 else None, zip(col_names[sorts[:, i]], corrs[sorts[:, i], i])))
        cols.append((col_mapping[i], col))
    return cols


def extract_diag_med_corres(D, Us, Ud, Um, dxmap_idx2desc, rxmap_idx2desc):
    corrs = {}
    for dx_idx in range(Ud.shape[0]):
        pts = D[:, dx_idx].nonzero()[0]
        pt_factor = Us[pts, :].sum(axis=0)
        corrs_pt = Um @ np.diag(pt_factor) @ Ud[dx_idx, :].reshape(-1, 1)
        corrs_pt = corrs_pt / (corrs_pt.sum()+1e-15)
        
        nnz_idx = corrs_pt.nonzero()[0]
        corrs_pt = [(rxmap_idx2desc[i], corrs_pt[i][0]) for i in nnz_idx]
        corrs_pt = sorted(corrs_pt, key=lambda x: x[1], reverse=True)
        
        corrs[dxmap_idx2desc[dx_idx]] = corrs_pt
    
    dx_count = D.sum(axis=0)
    dx_order = sorted(range(Ud.shape[0]), key=lambda x: dx_count[x], reverse=True)
    corrs = [(dxmap_idx2desc[i], corrs[dxmap_idx2desc[i]]) for i in dx_order]
    return corrs


def coord(i, j):
    return '{}{}'.format(get_column_letter(j), i)


def phenotypes_to_excel_worksheet(phenotypes, ws):
    n_dims = len(phenotypes[0])
    for i, pheno_r in enumerate(phenotypes):
        ws.merge_cells(coord(1, (n_dims + 1) * i + 1) + ':' + coord(1, (n_dims + 1) * i + n_dims))
        ws[coord(1, (n_dims + 1) * i + 1)] = 'Phenotype {:d}'.format(i + 1)
        ws[coord(1, (n_dims + 1) * i + 1)].font = Font(bold=True)
        ws[coord(1, (n_dims + 1) * i + 1)].alignment = Alignment(horizontal='center', vertical='center')

        for j, name in enumerate(phenotypes.dim_names):
            ws[coord(2, (n_dims + 1) * i + j + 1)] = name
            ws[coord(2, (n_dims + 1) * i + j + 1)].alignment = Alignment(horizontal='center', vertical='center')
            ws[coord(2, (n_dims + 1) * i + j + 1)].font = Font(bold=True)
            for k, (idx, item, weight) in enumerate(pheno_r[j]):
                ws[coord(k + 3, (n_dims + 1) * i + j + 1)] = '{} ({:.3f})'.format(item, weight)
                if phenotypes.stats['overlap'][j][idx] >= ceil(0.75 * len(phenotypes)):
                    ws.cell(row=k + 3, column=(n_dims + 1) * i + j + 1).fill = PatternFill(fgColor='FF8B8B',
                                                                                           fill_type='solid')
                elif phenotypes.stats['overlap'][j][idx] >= ceil(0.5 * len(phenotypes)):
                    ws.cell(row=k + 3, column=(n_dims + 1) * i + j + 1).fill = PatternFill(fgColor='FFBF8B',
                                                                                           fill_type='solid')
                elif phenotypes.stats['overlap'][j][idx] >= ceil(0.25 * len(phenotypes)):
                    ws.cell(row=k + 3, column=(n_dims + 1) * i + j + 1).fill = PatternFill(fgColor='FFFF97',
                                                                                           fill_type='solid')

            ws.column_dimensions[get_column_letter((n_dims + 1) * i + j + 1)].width = 50


def phenotypes_to_xlsx(phenotypes, filepath, ws_name=None):
    wb = Workbook()
    ws = wb.active
    if ws_name:
        ws.title = ws_name
    phenotypes_to_excel_worksheet(phenotypes, ws)
    wb.save(filepath)


def corrs_to_excel_worksheet(corrs_cols, ws):
    for dx_idx in range(len(corrs_cols)):
        col = corrs_cols[dx_idx]
        ws.merge_cells(coord(1, 3 * dx_idx + 1) + ':' + coord(1, 3 * dx_idx + 2))
        ws[coord(1, 3 * dx_idx + 1)] = col[0]
        ws[coord(1, 3 * dx_idx + 1)].font = Font(bold=True)
        ws[coord(1, 3 * dx_idx + 1)].alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(3 * dx_idx + 1)].width = 15
        ws.column_dimensions[get_column_letter(3 * dx_idx + 2)].width = 15
        ws.column_dimensions[get_column_letter(3 * dx_idx + 3)].width = 3

        for i, item in enumerate(col[1]):
            if item:
                ws[coord(i + 2, 3 * dx_idx + 1)] = item[0]
                ws[coord(i + 2, 3 * dx_idx + 2)] = item[1]
            else:
                break


def corrspondence_to_xlsx(corrs_cols, filepath, ws_name=None):
    wb = Workbook()
    ws = wb.active
    if ws_name:
        ws.title = ws_name
    corrs_to_excel_worksheet(corrs_cols, ws)
    wb.save(filepath)


def factor_matrices_to_xlsx(factors, item_dicts, filepath, threshold=None):
    dim_names, item_idx2desc = zip(*item_dicts)
    phenotypes = interpret_phenotypes(factors, item_idx2desc, dim_names, threshold=threshold)
    phenotypes_to_xlsx(phenotypes, filepath)


def batch_factors_to_xlsx(factors_dict, item_dicts, filepath, threshold=None):
    wb = Workbook()
    wb.remove(wb.active)

    dim_names, item_idx2desc = zip(*item_dicts)

    for name, factors in factors_dict.items():
        ws = wb.create_sheet(name)
        phenotypes = interpret_phenotypes(factors, item_idx2desc, dim_names, threshold=threshold)
        phenotypes_to_excel_worksheet(phenotypes, ws)
    wb.save(filepath)


def extract_corrs_to_xlsx(D, Us, Ud, Um, dxmap_idx2desc, rxmap_idx2desc, filepath, ws_name=None):
    corrs = extract_diag_med_corres(D, Us, Ud, Um, dxmap_idx2desc, rxmap_idx2desc)
    corrspondence_to_xlsx(corrs, filepath, ws_name=ws_name)